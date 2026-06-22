"""
competitors.py — Telehealth / home-care competitor config + xlsx reader.

Two jobs:
  1. TELEHEALTH_TERMS — the ~32 marketplace/telehealth/home-care competitors
     from competitor_matrix.xlsx, with curated Reddit search terms (generic
     names like "Heal", "Honor", "Papa", "Vitals" are disambiguated so we
     don't drown in unrelated hits).
  2. read_matrix() — parse competitor_matrix.xlsx into JSON for the dashboard's
     Competitor Matrix view (matrix + your-build-vs-threats + gap analysis).

Run `python competitors.py` to (re)generate data/competitor_matrix.json.
"""

import json
import os

CATEGORY_TELEHEALTH = "Telehealth/Home-Care"
CATEGORY_EHR = "EHR"

MATRIX_XLSX = "competitor_matrix.xlsx"
MATRIX_JSON = os.path.join("data", "competitor_matrix.json")

# Company -> curated Reddit/app search terms. Generic words disambiguated.
TELEHEALTH_TERMS = {
    "Zocdoc": ["Zocdoc"],
    "Healthgrades": ["Healthgrades"],
    "Vitals": ["Vitals.com doctor"],
    "Solv Health": ["Solv Health"],
    "Teladoc": ["Teladoc"],
    "MDLive": ["MDLive"],
    "Amwell": ["Amwell"],
    "Doctor on Demand": ["Doctor on Demand"],
    "HealthTap": ["HealthTap"],
    "PlushCare": ["PlushCare"],
    "Hims & Hers": ["Hims and Hers", "Hims & Hers"],
    "One Medical": ["One Medical"],
    "Forward Health": ["Forward Health primary care"],
    "Parsley Health": ["Parsley Health"],
    "Sesame": ["Sesame Care"],
    "DispatchHealth": ["DispatchHealth"],
    "MedArrive": ["MedArrive"],
    "Landmark Health": ["Landmark Health"],
    "Heal": ["Heal house call doctor"],
    "Included Health": ["Included Health"],
    "CareMore Health": ["CareMore Health"],
    "Contessa Health": ["Contessa Health"],
    "WellBe Senior": ["WellBe Senior Medical"],
    "Honor": ["Honor home care"],
    "Care.com": ["Care.com"],
    "Papa": ["Papa Pals senior"],
    "Home Instead": ["Home Instead"],
    "Visiting Angels": ["Visiting Angels"],
    "Bayada": ["Bayada home health"],
    "Hometeam": ["Hometeam home care"],
    "Amazon Pharmacy": ["Amazon Pharmacy"],
    "Capsule": ["Capsule pharmacy"],
    "Alto Pharmacy": ["Alto Pharmacy"],
}

TELEHEALTH_ORDER = list(TELEHEALTH_TERMS.keys())


def read_matrix():
    """Parse the xlsx into a JSON-able dict. Returns None if file/lib missing."""
    if not os.path.exists(MATRIX_XLSX):
        return None
    try:
        import openpyxl
    except ImportError:
        return None

    wb = openpyxl.load_workbook(MATRIX_XLSX, read_only=True, data_only=True)
    out = {"sheets": {}}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        # Drop fully-empty rows.
        rows = [r for r in rows if any(c not in (None, "") for c in r)]
        if not rows:
            continue
        # Heuristic: the header row is the first row with >3 non-empty cells
        # after the title rows. Title = single-cell rows at the top.
        header_idx = 0
        for i, r in enumerate(rows):
            nonempty = sum(1 for c in r if c not in (None, ""))
            if nonempty >= 3:
                header_idx = i
                break
        title = " / ".join(str(rows[0][0]) for _ in [0]) if header_idx > 0 else ""
        header = [str(c).replace("\n", " ").strip() if c is not None else ""
                  for c in rows[header_idx]]
        # Trim trailing empty header columns.
        while header and header[-1] == "":
            header.pop()
        ncol = len(header)
        data_rows = []
        for r in rows[header_idx + 1:]:
            vals = ["" if c is None else str(c) for c in r[:ncol]]
            if any(v.strip() for v in vals):
                data_rows.append(vals)
        out["sheets"][sheet_name] = {
            "title": title,
            "columns": header,
            "rows": data_rows,
        }
    return out


def main():
    matrix = read_matrix()
    if matrix is None:
        print(f"Could not read {MATRIX_XLSX} (missing file or openpyxl).")
        return
    os.makedirs("data", exist_ok=True)
    with open(MATRIX_JSON, "w", encoding="utf-8") as fh:
        json.dump(matrix, fh, ensure_ascii=False, indent=2)
    for name, sh in matrix["sheets"].items():
        print(f"  {name}: {len(sh['rows'])} rows x {len(sh['columns'])} cols")
    print(f"Saved {MATRIX_JSON}")


if __name__ == "__main__":
    main()
